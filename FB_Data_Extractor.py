import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows  

def verificar_link():
    link = entry_link.get()
    num_datos_obtenidos = int(entry_num_datos.get())  
    if "facebook.com" in link:
        messagebox.showinfo("¡Enlace válido!", "El enlace es de Facebook.")
        abrir_navegador(link, num_datos_obtenidos)
        root.destroy()  
    else:
        messagebox.showerror("Error", "El enlace no es de Facebook.")

def ayuda(): 
    messagebox.showinfo("Contacto", "21030503@upgto.edu.mx o tel. 2288549203moni.")

def cerrar_ventana():
    root.destroy()

def abrir_navegador(link, num_datos_obtenidos):
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--disable-extensions')
    driver = webdriver.Chrome(options=options)
    driver.set_window_position(2000, 0)
    driver.maximize_window()
    
    try:
        driver.get(link)
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[aria-label="Cerrar"] i')))
        driver.find_element(By.CSS_SELECTOR, 'div[aria-label="Cerrar"] i').click()
        time.sleep(40)
        
   
        driver.execute_script("window.scrollTo(2, document.body.scrollHeight);")
        time.sleep(2)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        elementos = soup.find_all('span', class_=re.compile('^x1e558r4$'))

        datos_numericos = []
        palabras_clave = []
        comentarios = []

        for i, elemento in enumerate(elementos):
            if len(datos_numericos) < num_datos_obtenidos:  
                texto = elemento.text.strip()  
                if texto: 
                    datos_numericos.append(texto)
            else:
                break

        
        elementos_a = soup.find_all('a', attrs={'aria-label': re.compile(r'.*')})
        
        for elemento_a in elementos_a:
            if len(palabras_clave) < num_datos_obtenidos:
                span_elemento = elemento_a.find('span')
                if span_elemento: 
                    palabra_clave = span_elemento.text.strip()
                    palabras_clave.append(palabra_clave)
            else:
                break

       
        elementos_b = soup.find_all('div', attrs={'data-testid': re.compile(r'.*')})
        for elemento_b in elementos_b[1:]: 
            if len(comentarios) < num_datos_obtenidos:
                span_elemento = elemento_b.find('span')
                if span_elemento: 
                    comentario = span_elemento.text.strip()
                    comentarios.append(comentario)
            else:
                break

        if datos_numericos:
            df_numericos = pd.DataFrame({'Datos': datos_numericos})
            df_numericos['Datos'] = df_numericos['Datos'].str.strip()  
            print("Datos numéricos encontrados:")
            print(df_numericos)

        if palabras_clave:
            df_palabras_clave = pd.DataFrame({'Palabras Clave': palabras_clave})
            print("\nPalabras Clave encontradas:")
            print(df_palabras_clave)

        guardar_en_excel(df_numericos, df_palabras_clave)

    except Exception as e:
        print("Se ha producido un error:", e)
    finally:
        driver.quit()

def guardar_en_excel(df_numericos, df_palabras_clave):
  
    wb = Workbook()
    ws = wb.active

    if not df_numericos.empty:
        ws1 = wb.create_sheet("Numero de reacciones")
        for r in dataframe_to_rows(df_numericos, index=False, header=True):
            ws1.append(r)

    if not df_palabras_clave.empty:
        ws2 = wb.create_sheet("Fechas")
        for r in dataframe_to_rows(df_palabras_clave, index=False, header=True):
            ws2.append(r)

    wb.save("resultados_facebook.xlsx")
    print("Los resultados se han guardado en 'resultados_facebook.xlsx'")

root = tk.Tk()
root.title("Verificar Enlace de Facebook")

image = Image.open("D:/Darcknek/imagen/Captura.jpg")  
background_image = ImageTk.PhotoImage(image)

canvas = tk.Canvas(root, width=image.width, height=image.height)
canvas.pack(fill="both", expand=True)
canvas.create_image(0, 0, image=background_image, anchor="nw")

label_link = tk.Label(root, text="Ingresa el enlace de Facebook:", bg="white")
label_link.place(relx=0.5, rely=0.3, anchor="center")
entry_link = tk.Entry(root, width=50)
entry_link.place(relx=0.5, rely=0.37, anchor="center")
label_num_datos = tk.Label(root, text="Número de datos requeridos:", bg="white")
label_num_datos.place(relx=0.5, rely=0.43, anchor="center")
entry_num_datos = tk.Entry(root, width=10)
entry_num_datos.place(relx=0.5, rely=0.49, anchor="center")
button_verificar = tk.Button(root, text="Verificar Enlace", command=verificar_link)
button_verificar.place(relx=0.3, rely=0.62, anchor="center")
button_cerrar = tk.Button(root, text="Cerrar Ventana", command=cerrar_ventana)
button_cerrar.place(relx=0.7, rely=0.62, anchor="center")
button_ayuda = tk.Button(root, text="Ayuda", command=ayuda)
button_ayuda.place(relx=0.5, rely=0.62, anchor="center")

root.mainloop()
